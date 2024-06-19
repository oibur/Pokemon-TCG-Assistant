import openpyxl

def add_concatenated_column_and_insert_empty_columns(cards_filename, my_cards_filename):
    # Load the CARDS workbook
    cards_wb = openpyxl.load_workbook(cards_filename)

    # Load the MY.CARDS workbook
    my_cards_wb = openpyxl.load_workbook(my_cards_filename)
    my_cards_sheet = my_cards_wb.active

    # Create a dictionary for quick lookup of MY.CARDS data
    my_cards_data = {}
    for row in range(2, my_cards_sheet.max_row + 1):
        key = my_cards_sheet.cell(row=row, column=1).value  # Column A
        if key:
            my_cards_data[key] = {my_cards_sheet.cell(row=1, column=col).value: my_cards_sheet.cell(row=row, column=col).value for col in range(2, my_cards_sheet.max_column + 1)}

    # Loop through all sheets except the first one in CARDS.xlsx
    for sheet_name in cards_wb.sheetnames[1:]:
        cards_sheet = cards_wb[sheet_name]
        
        # Insert a new column between A and B and concatenate values from B and C
        cards_sheet.insert_cols(2)
        
        for row in range(2, cards_sheet.max_row + 1):
            cell_b = cards_sheet.cell(row=row, column=3).value  # Column B (now column C)
            cell_c = cards_sheet.cell(row=row, column=4).value  # Column C (now column D)
            concatenated_value = f"{cell_b} {cell_c}" if cell_b and cell_c else ''
            cards_sheet.cell(row=row, column=2).value = concatenated_value  # New column B
        
        # Iterate through columns G and onward in CARDS.xlsx
        col_index = 7  # Column G is the 7th column
        while col_index <= cards_sheet.max_column:
            cell = cards_sheet.cell(row=1, column=col_index)
            if cell.value and isinstance(cell.value, str):
                # Insert two empty columns after the current column
                cards_sheet.insert_cols(col_index + 1, amount=2)
                
                # Name the newly inserted columns
                cards_sheet.cell(row=1, column=col_index + 1).value = '#'
                cards_sheet.cell(row=1, column=col_index + 2).value = '$'
                
                # Move to the next text column after the inserted columns
                col_index += 3
            else:
                col_index += 1

        # Now populate '#' columns with data from MY.CARDS.xlsx
        for col in range(7, cards_sheet.max_column + 1):
            header_cell = cards_sheet.cell(row=1, column=col)
            if header_cell.value == '#':
                previous_col_header = cards_sheet.cell(row=1, column=col - 1).value
                if previous_col_header:
                    for row in range(2, cards_sheet.max_row + 1):
                        cards_key = cards_sheet.cell(row=row, column=3).value  # Matching column B (now C) in CARDS.xlsx
                        if cards_key and cards_key in my_cards_data:
                            cards_sheet.cell(row=row, column=col).value = my_cards_data[cards_key].get(previous_col_header)

    # Save the workbook with the changes
    cards_wb.save(cards_filename)

# Call the function with the filenames
add_concatenated_column_and_insert_empty_columns('CARDS.xlsx', 'MY.CARDS.xlsx')
